import pandas as pd
from sentence_transformers import SentenceTransformer
from openpyxl import Workbook
from openpyxl.styles import Font, Color, colors, fills
from openpyxl.utils.dataframe import dataframe_to_rows
wb = Workbook()
ws = wb.active


model_sentence_transformers = SentenceTransformer('./model/labse_bert_model')


df = pd.read_excel('./data/Job 220138 - Chinese - Milestone 4_.xlsx',
                   header=None)

for i, r in enumerate(dataframe_to_rows(df, index=False, header=False)):
    ws.append(r)
    embeddings = model_sentence_transformers.encode(r,
                                                    show_progress_bar=False,
                                                    convert_to_numpy=True,
                                                    normalize_embeddings=True)

    cos1 = embeddings[1].dot(embeddings[2])
    cos2 = embeddings[1].dot(embeddings[3])

    ws['E'+str(i+1)] = cos1
    ws['F'+str(i+1)] = cos2
    ws['G'+str(i+1)] = cos2-cos1


wb.save('./data/Job 220138 - Chinese - Milestone 4_.xlsx')
print('finished')